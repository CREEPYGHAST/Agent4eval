#!/usr/bin/env python3
import argparse
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Tuple

import cv2
import numpy as np
import pytesseract
from openpyxl import Workbook


FILENAME_PATTERN = re.compile(r"^(?P<score>\d+)_(?P<name>.+)_(?P<student_id>\d+)\((?P<page>[12])\)\.jpg$", re.IGNORECASE)


@dataclass(frozen=True)
class StudentKey:
    score: int
    name: str
    student_id: str


@dataclass
class Config:
    suffix_mask_width_ratio: float = 0.25
    suffix_mask_height_ratio: float = 0.25
    red_mask_min_area: int = 120
    red_mask_dilate_kernel: int = 5
    adaptive_block_size: int = 31
    adaptive_c: int = 12


def parse_filename(path: Path) -> Tuple[StudentKey, int]:
    match = FILENAME_PATTERN.match(path.name)
    if not match:
        raise ValueError(f"Invalid filename format: {path.name}")

    key = StudentKey(
        score=int(match.group("score")),
        name=match.group("name").strip(),
        student_id=match.group("student_id"),
    )
    page = int(match.group("page"))
    return key, page


def load_image(path: Path) -> np.ndarray:
    data = np.fromfile(str(path), dtype=np.uint8)
    image = cv2.imdecode(data, cv2.IMREAD_COLOR)
    if image is None:
        raise ValueError(f"Unable to read image: {path}")
    return image


def save_image(path: Path, image: np.ndarray) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    ext = path.suffix if path.suffix else ".jpg"
    ok, encoded = cv2.imencode(ext, image)
    if not ok:
        raise ValueError(f"Unable to encode image for saving: {path}")
    encoded.tofile(str(path))


def pad_to_width(image: np.ndarray, target_width: int) -> np.ndarray:
    h, w = image.shape[:2]
    if w == target_width:
        return image
    pad_left = (target_width - w) // 2
    pad_right = target_width - w - pad_left
    return cv2.copyMakeBorder(
        image,
        top=0,
        bottom=0,
        left=pad_left,
        right=pad_right,
        borderType=cv2.BORDER_CONSTANT,
        value=(255, 255, 255),
    )


def stitch_images(images: List[np.ndarray]) -> np.ndarray:
    if len(images) == 1:
        return images[0]

    max_width = max(im.shape[1] for im in images)
    aligned = [pad_to_width(im, max_width) for im in images]
    return cv2.vconcat(aligned)


def mask_red_annotations(image: np.ndarray, config: Config) -> np.ndarray:
    hsv = cv2.cvtColor(image, cv2.COLOR_BGR2HSV)

    lower_red_1 = np.array([0, 70, 70])
    upper_red_1 = np.array([10, 255, 255])
    lower_red_2 = np.array([160, 70, 70])
    upper_red_2 = np.array([180, 255, 255])

    mask1 = cv2.inRange(hsv, lower_red_1, upper_red_1)
    mask2 = cv2.inRange(hsv, lower_red_2, upper_red_2)
    red_mask = cv2.bitwise_or(mask1, mask2)

    kernel = np.ones((config.red_mask_dilate_kernel, config.red_mask_dilate_kernel), np.uint8)
    red_mask = cv2.morphologyEx(red_mask, cv2.MORPH_CLOSE, kernel)
    red_mask = cv2.dilate(red_mask, kernel, iterations=1)

    contours, _ = cv2.findContours(red_mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    filtered_mask = np.zeros_like(red_mask)
    for cnt in contours:
        if cv2.contourArea(cnt) >= config.red_mask_min_area:
            cv2.drawContours(filtered_mask, [cnt], -1, 255, thickness=cv2.FILLED)

    output = image.copy()
    output[filtered_mask > 0] = (255, 255, 255)
    return output


def mask_suffix_region(image: np.ndarray, config: Config) -> np.ndarray:
    h, w = image.shape[:2]
    mask_w = int(w * config.suffix_mask_width_ratio)
    mask_h = int(h * config.suffix_mask_height_ratio)

    output = image.copy()
    x1 = max(w - mask_w, 0)
    y1 = max(h - mask_h, 0)
    output[y1:h, x1:w] = (255, 255, 255)
    return output


def preprocess_for_ocr(image: np.ndarray, config: Config) -> np.ndarray:
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    denoised = cv2.fastNlMeansDenoising(gray, None, h=8, templateWindowSize=7, searchWindowSize=21)
    clahe = cv2.createCLAHE(clipLimit=2.5, tileGridSize=(8, 8))
    enhanced = clahe.apply(denoised)

    block_size = config.adaptive_block_size
    if block_size % 2 == 0:
        block_size += 1
    block_size = max(3, block_size)

    binary = cv2.adaptiveThreshold(
        enhanced,
        255,
        cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
        cv2.THRESH_BINARY,
        block_size,
        config.adaptive_c,
    )
    return binary


def clean_text(text: str, student_name: str) -> str:
    lines = [ln.strip() for ln in text.splitlines()]
    cleaned: List[str] = []
    score_pattern = re.compile(r"\(\d+\s*/\s*\d+\s*分\)")

    for line in lines:
        if not line:
            continue
        if student_name and student_name in line:
            continue
        if score_pattern.search(line):
            continue
        if re.search(r"yours\s*,?\s*li\s*hua", line, re.IGNORECASE):
            continue
        cleaned.append(line)

    return "\n".join(cleaned).strip()


def run_ocr(image: np.ndarray, lang: str) -> str:
    config = "--oem 3 --psm 6"
    return pytesseract.image_to_string(image, lang=lang, config=config)


def process_student(
    key: StudentKey,
    page_images: List[Tuple[int, Path]],
    config: Config,
    ocr_lang: str,
    debug_dir: Path | None,
) -> Tuple[StudentKey, str]:
    sorted_pages = sorted(page_images, key=lambda x: x[0])
    images = [load_image(path) for _, path in sorted_pages]

    stitched = stitch_images(images)
    red_masked = mask_red_annotations(stitched, config)
    suffix_masked = mask_suffix_region(red_masked, config)
    preprocessed = preprocess_for_ocr(suffix_masked, config)

    raw_text = run_ocr(preprocessed, lang=ocr_lang)
    final_text = clean_text(raw_text, key.name)

    if debug_dir:
        student_stub = f"{key.score}_{key.name}_{key.student_id}"
        save_image(debug_dir / f"{student_stub}_stitched.jpg", stitched)
        save_image(debug_dir / f"{student_stub}_masked.jpg", suffix_masked)
        save_image(debug_dir / f"{student_stub}_ocr_ready.jpg", preprocessed)

    return key, final_text


def write_excel(results: List[Tuple[StudentKey, str]], output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "results"
    ws.append(["原始得分", "姓名", "学号", "提取文本"])

    for key, text in results:
        ws.append([key.score, key.name, key.student_id, text])

    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=3).number_format = "@"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def collect_groups(input_dir: Path) -> Dict[StudentKey, List[Tuple[int, Path]]]:
    groups: Dict[StudentKey, List[Tuple[int, Path]]] = {}

    for path in sorted(input_dir.glob("*.jpg")):
        key, page = parse_filename(path)
        groups.setdefault(key, []).append((page, path))

    for key, items in groups.items():
        pages = [p for p, _ in items]
        if len(set(pages)) != len(pages):
            raise ValueError(f"Duplicate page index found for: {key}")
        if any(p not in (1, 2) for p in pages):
            raise ValueError(f"Unsupported page index found for: {key}")

    return groups


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="作文答题卡图片批处理（拼接 + 去干扰 + OCR）并导出 Excel")
    parser.add_argument("input_dir", type=Path, help="包含 .jpg 图片的输入目录")
    parser.add_argument("--output", type=Path, default=Path("results.xlsx"), help="输出 xlsx 路径")
    parser.add_argument("--ocr-lang", default="chi_sim+eng", help="pytesseract OCR 语言包")
    parser.add_argument("--suffix-mask-width-ratio", type=float, default=0.25, help="右下角遮罩宽度比例")
    parser.add_argument("--suffix-mask-height-ratio", type=float, default=0.25, help="右下角遮罩高度比例")
    parser.add_argument("--debug-dir", type=Path, default=None, help="可选：输出中间图像目录")
    return parser


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()

    if not args.input_dir.exists() or not args.input_dir.is_dir():
        raise SystemExit(f"Input directory not found: {args.input_dir}")

    config = Config(
        suffix_mask_width_ratio=args.suffix_mask_width_ratio,
        suffix_mask_height_ratio=args.suffix_mask_height_ratio,
    )

    groups = collect_groups(args.input_dir)
    results: List[Tuple[StudentKey, str]] = []

    for key in sorted(groups.keys(), key=lambda x: (x.name, x.student_id, x.score)):
        results.append(process_student(key, groups[key], config, args.ocr_lang, args.debug_dir))

    write_excel(results, args.output)
    print(f"Processed {len(results)} students. Output: {args.output}")


if __name__ == "__main__":
    main()
